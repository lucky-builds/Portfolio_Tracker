from io import BytesIO
from typing import Optional

import msoffcrypto
import openpyxl
import requests


class AngelOneParseError(Exception):
    pass


def parse_angelone_xlsx(file_bytes: bytes, password: Optional[str] = None) -> dict:
    """
    Parse an Angel One 'Your Holding Details' XLSX report.

    Returns:
        {
            "client_name": str,
            "client_id": str,
            "equity": [{"company_name", "isin", "qty", "avg_price"}],
            "mf": [{"fund_name", "isin", "units", "avg_nav", "suggested_scheme_code"}],
        }
    """
    try:
        buf = BytesIO(file_bytes)
        if password:
            office_file = msoffcrypto.OfficeFile(buf)
            office_file.load_key(password=password)
            decrypted = BytesIO()
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            wb = openpyxl.load_workbook(decrypted, data_only=True)
        else:
            wb = openpyxl.load_workbook(buf, data_only=True)
    except Exception as e:
        raise AngelOneParseError(f"Failed to open file: {e}")

    client_name, client_id = _extract_client_info(wb)
    equity = _parse_equity_sheet(wb)
    mf = _parse_mf_sheet(wb)

    return {
        "client_name": client_name,
        "client_id": client_id,
        "equity": equity,
        "mf": mf,
    }


def _extract_client_info(wb) -> tuple[str, str]:
    sheet_name = next((s for s in ["Summary", "Equity"] if s in wb.sheetnames), None)
    if not sheet_name:
        return "", ""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        if row and row[0] == "Client Name" and i + 1 < len(rows):
            data_row = rows[i + 1]
            return str(data_row[0] or "").strip(), str(data_row[1] or "").strip()
    return "", ""


def _parse_equity_sheet(wb) -> list:
    if "Equity" not in wb.sheetnames:
        return []
    ws = wb["Equity"]
    rows = list(ws.iter_rows(values_only=True))

    # Locate the detail header row (contains "Company Name" at index 1)
    header_idx = None
    for i, row in enumerate(rows):
        if row and row[1] == "Company Name":
            header_idx = i
            break
    if header_idx is None:
        return []

    holdings = []
    for row in rows[header_idx + 1:]:
        if not row or not row[0]:
            continue
        if str(row[0]).strip() == "Total":
            break

        company_name = str(row[1] or "").strip()
        isin = str(row[2] or "").strip()
        total_qty = row[5]        # Total Quantity
        avg_price = row[12]       # Avg Trading Price

        if not company_name or total_qty is None or avg_price is None:
            continue
        try:
            qty = int(float(str(total_qty)))
            avg_price_val = round(float(str(avg_price)), 2)
        except (ValueError, TypeError):
            continue
        if qty <= 0:
            continue

        holdings.append({
            "company_name": company_name,
            "isin": isin,
            "qty": qty,
            "avg_price": avg_price_val,
        })
    return holdings


def _parse_mf_sheet(wb) -> list:
    if "Mutual Funds" not in wb.sheetnames:
        return []
    ws = wb["Mutual Funds"]
    rows = list(ws.iter_rows(values_only=True))

    # Locate the detail header row (contains "Fund Name" at index 1)
    header_idx = None
    for i, row in enumerate(rows):
        if row and row[1] == "Fund Name":
            header_idx = i
            break
    if header_idx is None:
        return []

    holdings = []
    for row in rows[header_idx + 1:]:
        if not row or not row[0]:
            continue
        if str(row[0]).strip() == "Total":
            break

        fund_name = str(row[1] or "").strip()
        isin = str(row[2] or "").strip()
        units = row[3]      # Units
        avg_nav = row[4]    # Average NAV

        if not fund_name or units is None or avg_nav is None:
            continue
        try:
            units_val = round(float(str(units)), 3)
            avg_nav_val = round(float(str(avg_nav)), 4)
        except (ValueError, TypeError):
            continue
        if units_val <= 0:
            continue

        scheme_code = _resolve_mf_scheme_code(fund_name)
        holdings.append({
            "fund_name": fund_name,
            "isin": isin,
            "units": units_val,
            "avg_nav": avg_nav_val,
            "suggested_scheme_code": scheme_code,
        })
    return holdings


def _resolve_mf_scheme_code(fund_name: str) -> Optional[str]:
    """Try to find a matching scheme code from mfapi.in search."""
    try:
        resp = requests.get(
            "https://api.mfapi.in/mf/search",
            params={"q": fund_name},
            timeout=5,
        )
        if resp.ok:
            results = resp.json()
            if results:
                return str(results[0]["schemeCode"])
    except Exception:
        pass
    return None
